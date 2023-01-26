
import openpyxl

file = openpyxl.load_workbook('inventory.xlsx')


product_details = file['Sheet1']

products_company = {}
products_cost_per_company = {}


for product_row in range(2, product_details.max_row + 1):
    # Total Company
    company_name = product_details.cell(product_row, 4).value
    if company_name in products_company:
        products_company[company_name] = products_company.get(company_name) + 1
    else:
        products_company[company_name] = 1

    # Total Cost per Company
    product_quantity = product_details.cell(product_row, 2).value
    product_price = product_details.cell(product_row, 3).value

    if company_name in products_cost_per_company:
        products_cost_per_company[company_name] = products_cost_per_company.get(company_name) +\
                                                  (product_quantity * product_price)
    else:
        products_cost_per_company[company_name] = product_quantity * product_price
    # Calculate Unit Price and Save to Excel Sheet
    unit_price = product_details.cell(product_row, 5)
    unit_price.value = round(product_price / product_quantity, 2)

    # Calculate Total Cost per company and save to Excel Sheet
    total_cost = product_details.cell(product_row, 6)
    total_cost.value = round(product_price * product_quantity, 2)


e5 = product_details.cell(1, 5)
e6 = product_details.cell(1, 6)
e6.value = " Total Cost"
e5.value = 'Unit Price'


file.save("inventory_solution_unitPrice_totalPrice.xlsx")

print(products_company)
print(products_cost_per_company)
print(sum(products_cost_per_company.values()))



